"""Core CDD -> RETConfigWDTInternal conversion logic (no UI).

Shared by generate_ret.py (CLI) and ret_gui.py (Tkinter GUI).
"""
import json
import re
from collections import defaultdict
from copy import copy

from openpyxl import load_workbook

# 4 devices per sector: order matches positions 1..4 in the template.
DEVICE_ORDER = ["C", "D", "E", "_NOT_USED_"]

# Output column headers (target order).
HEADERS = [
    "Site Name",
    "RRU Name",
    "RRU CN",
    "RRU SRN",
    "RRU SN",
    "RCU Coloring",
    "RCU Tilt",
    "Device Name",
]


def load_mapping(path):
    with open(path, encoding="utf-8") as f:
        return json.load(f)


def _scaled_tilt(present, band_letter, fallback=None):
    """E_TILT (degrees) for a band -> 0.1-degree units, with an optional
    fallback band when the primary band has no E_TILT in this sector."""
    t = present.get(band_letter)
    if t is None and fallback is not None:
        t = present.get(fallback)
    return round((t or 0.0) * 10)


def _norm_header(s):
    """Normalize a header cell for tolerant matching (case/whitespace-insensitive)."""
    if s is None:
        return ""
    return " ".join(str(s).split()).strip().lower()


def _col_letter_to_index(letter):
    """'A' -> 0, 'B' -> 1, ... (single-letter columns)."""
    return ord(str(letter).strip().upper()[0]) - ord("A")


def _norm_uarfcn(v):
    """Normalize a UARFCN for tolerant equality (10612 == 10612.0 == '10612')."""
    if v is None:
        return ""
    try:
        return str(int(float(v)))
    except (TypeError, ValueError):
        return str(v).strip()


def _find_index(norm_headers, targets):
    """Index of the first header matching any target (exact, then startswith)."""
    targets = [t for t in targets if t]
    for t in targets:
        if t in norm_headers:
            return norm_headers.index(t)
    for i, h in enumerate(norm_headers):
        if h and any(h.startswith(t) for t in targets):
            return i
    return None


def _locate_header(all_rows, key_targets, max_scan=10):
    """Find the header row by locating the key column; return (row_index, norm_headers).

    Scans the first ``max_scan`` rows so a leading title/blank row doesn't break
    detection. Falls back to row 0 if the key column is never found.
    """
    for i, row in enumerate(all_rows[:max_scan]):
        norm = [_norm_header(h) for h in row]
        if _find_index(norm, key_targets) is not None:
            return i, norm
    return 0, [_norm_header(h) for h in (all_rows[0] if all_rows else ())]


def list_sheets(cdd_path):
    """Return sheet names in the CDD workbook."""
    wb = load_workbook(cdd_path, read_only=True)
    try:
        return wb.sheetnames
    finally:
        wb.close()


def load_three_g_tilts(cdd_wb, mapping):
    """Map (site_old, sector_number) -> New E-Tilt from the 3G sheet.

    The 3G Installation Design sheet lists a 'Logical Sector Name' formatted as
    '{SiteName_Old}-{sector number}' (e.g. 'KGPQ01-2'). Device No 3 (band E) is
    emitted as NSN_U2100 only when its sector appears here (else NOT_USED), and
    its tilt is taken from the 'New E-Tilt' column. A sector can span multiple
    rows (carriers); the first non-blank New E-Tilt wins, except a nonzero value
    overrides a previously stored 0 when three_g.etilt_prefer_nonzero (default
    true), so an unfilled (0) carrier row does not mask the sector's real tilt.
    Membership in the dict means the sector is present. Empty dict if the
    sheet/config is absent.
    """
    cfg = mapping.get("three_g")
    if not cfg or cfg["sheet"] not in cdd_wb.sheetnames:
        return {}
    ws = cdd_wb[cfg["sheet"]]
    all_rows = list(ws.iter_rows(min_row=1, values_only=True))
    if not all_rows:
        return {}

    # Locate columns by header name. The site MATCH key comes from the 'Site'
    # column (== 4G SiteName_Old); only the sector number comes from the
    # Logical Sector Name (text after the last separator).
    site_targets = [_norm_header(cfg.get("site_header", "Site"))]
    sec_targets = [_norm_header(cfg.get("logical_sector_header", "")),
                   _norm_header(cfg.get("logical_sector_header_prefix", "Logical Sector Name"))]
    tilt_targets = [_norm_header(cfg.get("etilt_header", "New E-Tilt"))]
    uarfcn_targets = [_norm_header(cfg.get("uarfcn_header", "UARFCN"))]
    hdr_idx, norm = _locate_header(all_rows, sec_targets)
    site_col = _find_index(norm, site_targets)
    sec_col = _find_index(norm, sec_targets)
    tilt_col = _find_index(norm, tilt_targets)
    uarfcn_col = _find_index(norm, uarfcn_targets)
    # Fall back to configured column letters if a header isn't found.
    if site_col is None and cfg.get("site_column"):
        site_col = _col_letter_to_index(cfg["site_column"])
    if sec_col is None and cfg.get("logical_sector_column"):
        sec_col = _col_letter_to_index(cfg["logical_sector_column"])
    if tilt_col is None and cfg.get("etilt_column"):
        tilt_col = _col_letter_to_index(cfg["etilt_column"])
    if uarfcn_col is None and cfg.get("uarfcn_column"):
        uarfcn_col = _col_letter_to_index(cfg["uarfcn_column"])
    if sec_col is None:
        return {}

    # The New E-Tilt is taken only from the carrier row whose UARFCN matches
    # three_g.etilt_uarfcn (10612 for CRb3). Other carrier rows still mark the
    # sector as present but contribute no tilt. Configurable in mapping.json.
    etilt_uarfcn = cfg.get("etilt_uarfcn")
    etilt_uarfcn_key = _norm_uarfcn(etilt_uarfcn) if etilt_uarfcn is not None else None

    sep = cfg.get("separator", "-")
    # When a sector spans multiple carrier rows, prefer a nonzero New E-Tilt over
    # a leading 0 (an unfilled carrier row should not mask the real tilt). When
    # false, the first non-blank value wins even if it is 0. Configurable in
    # mapping.json -> three_g.etilt_prefer_nonzero.
    prefer_nonzero = cfg.get("etilt_prefer_nonzero", True)
    tilts = {}
    for r in all_rows[hdr_idx + 1:]:
        if not r or len(r) <= sec_col or r[sec_col] is None:
            continue
        # Sector number = last segment of the Logical Sector Name (rpartition
        # so a site name containing the separator doesn't confuse it).
        _, _, num = str(r[sec_col]).rpartition(sep)
        if not num.strip():
            continue
        # Site comes from the 'Site' column when available, else the LSN prefix.
        if site_col is not None and len(r) > site_col and r[site_col] is not None:
            site_old = str(r[site_col])
        else:
            site_old = str(r[sec_col]).rpartition(sep)[0]
        key = (site_old.strip(), num.strip())
        e_tilt = r[tilt_col] if (tilt_col is not None and len(r) > tilt_col) else None
        try:
            e_tilt_val = float(e_tilt) if e_tilt not in (None, "-", "") else None
        except (TypeError, ValueError):
            e_tilt_val = None
        # Restrict the tilt source to the UARFCN==etilt_uarfcn carrier row. Rows
        # with any other UARFCN still register the sector (membership) but supply
        # no tilt value, so the 10612 carrier's New E-Tilt is the one that wins.
        if etilt_uarfcn_key is not None:
            row_uarfcn = (r[uarfcn_col]
                          if (uarfcn_col is not None and len(r) > uarfcn_col)
                          else None)
            if _norm_uarfcn(row_uarfcn) != etilt_uarfcn_key:
                e_tilt_val = None
        # First non-blank tilt wins; still register the key if only blanks seen.
        # With prefer_nonzero, a nonzero value also overrides a previously stored
        # 0 so an unfilled (0) carrier row doesn't mask the sector's real tilt.
        if key not in tilts:
            tilts[key] = e_tilt_val
        elif tilts[key] is None and e_tilt_val is not None:
            tilts[key] = e_tilt_val
        elif (prefer_nonzero and tilts[key] in (None, 0.0)
                and e_tilt_val not in (None, 0.0)):
            tilts[key] = e_tilt_val
    return tilts


def build_rows(cdd_path, sheet, mapping):
    """Parse the CDD sheet and produce output rows.

    Returns (rows, skipped, sector_count) where:
      rows      = list of 8-value lists (matching HEADERS)
      skipped   = list of (cell_name, reason)
      sector_count = number of distinct sectors emitted
    """
    band_map = mapping["band_map"]
    sector_map = mapping["sector_map"]
    l1800_naming = mapping.get("l1800_device_naming")
    cell_id_map = mapping.get("cell_id_map", {})
    band_by_tens = cell_id_map.get("band_by_tens", {})
    sector_by_units = cell_id_map.get("sector_by_units", {})
    rru_cn = mapping["constants"]["rru_cn"]
    rru_sn = mapping["constants"]["rru_sn"]
    three_g = mapping.get("three_g", {})
    tg_band = three_g.get("governs_band")            # e.g. "E"
    tg_absent = three_g.get("absent_band_token", "NOT_USED")
    # Case 8: site is in the 3G sheet but this sector isn't (4G has more sectors
    # than 3G). Both this and a wholly-absent site emit U2100_NOT_USED per
    # mapping.json; kept separate so the tokens can diverge again if needed.
    tg_partial_absent = three_g.get("partial_absent_band_token", tg_absent)
    # RRU Name prefix name: NEName_New (ne_name) by default, falling back to
    # SiteName_New (site_new) when blank. Configurable in field_rules.rru_name.
    rru_cfg = mapping.get("field_rules", {}).get("rru_name", {})
    prefix_source = rru_cfg.get("prefix_source", "site_new")
    prefix_fallback = rru_cfg.get("prefix_fallback", "site_new")

    cdd_wb = load_workbook(cdd_path, data_only=True)
    cdd_ws = cdd_wb[sheet]
    all_rows = list(cdd_ws.iter_rows(min_row=1, values_only=True))
    three_g_tilts = load_three_g_tilts(cdd_wb, mapping)
    # Sites that appear in the 3G sheet at all (any sector). Used to distinguish
    # Case 8 (site present, this sector missing) from a site wholly absent.
    three_g_sites = {site_old for (site_old, _num) in three_g_tilts}
    cdd_wb.close()

    # Resolve columns by HEADER NAME so reordered/extra columns don't break parsing.
    hdr_cfg = mapping.get("source", {}).get("headers", {})
    wanted = {
        "site_old": hdr_cfg.get("site_old", "SiteName (RRU Location)_Old"),
        "site_new": hdr_cfg.get("site_new", "SiteName (RRU Location)_New"),
        "ne_name": hdr_cfg.get("ne_name", "NEName_New"),
        "ne_id": hdr_cfg.get("ne_id", "Ne ID (New)"),
        "cell_name": hdr_cfg.get("cell_name", "CellName (New)[Key]"),
        "e_tilt": hdr_cfg.get("e_tilt", "E_TILT"),
        "local_cell_id": hdr_cfg.get("local_cell_id", "Local Cell ID (New)"),
        "tx_rx_mode": hdr_cfg.get("tx_rx_mode", "TxRxMode (New)"),
    }
    cell_targets = [_norm_header(wanted["cell_name"])]
    hdr_idx, norm = _locate_header(all_rows, cell_targets)
    col = {f: _find_index(norm, [_norm_header(name)]) for f, name in wanted.items()}

    required = ("site_old", "site_new", "ne_id", "cell_name")
    missing = [wanted[f] for f in required if col[f] is None]
    if missing:
        found = [str(h) for h in (all_rows[hdr_idx] if all_rows else ()) if h is not None]
        raise ValueError(
            "Sheet '%s': could not find required column(s) by header name: %s.\n"
            "Headers found in the sheet:\n  %s\n"
            "Check the column names in mapping.json -> source.headers."
            % (sheet, ", ".join('\"%s\"' % m for m in missing), "\n  ".join(found))
        )

    src_rows = all_rows[hdr_idx + 1:]
    by_sector = defaultdict(dict)   # (site_new, ne_id, Y) -> {X: e_tilt}
    sector_site_old = {}            # (site_new, ne_id, Y) -> site_old
    site_old_by_site = {}           # (site_new, ne_id) -> site_old (site-wide)
    sector_ne_name = {}             # (site_new, ne_id, Y) -> NEName_New
    by_sector_txrx = defaultdict(dict)  # (site_new, ne_id, Y) -> {X: TxRxMode}
    sector_order = []
    seen_sector = set()
    skipped = []

    def _get(r, field):
        j = col[field]
        return r[j] if (j is not None and len(r) > j) else None

    for r in src_rows:
        if not r or _get(r, "cell_name") is None:
            continue
        site_old = _get(r, "site_old")
        site_old = str(site_old).strip() if site_old is not None else ""
        site_new = _get(r, "site_new")
        ne_id = _get(r, "ne_id")
        cell = str(_get(r, "cell_name"))
        e_tilt = _get(r, "e_tilt")
        if len(cell) < 2:
            skipped.append((cell, "too short"))
            continue
        # Band (X) and sector (Y): Local Cell ID (New) is authoritative when
        # present (tens digit -> band, units digit -> sector); otherwise fall
        # back to the CellName tail chars (cell[-2], cell[-1]).
        X = Y = None
        lcid = _get(r, "local_cell_id")
        if lcid is not None:
            digits = str(lcid).strip()
            if digits.isdigit() and len(digits) >= 2:
                X = band_by_tens.get(digits[-2])
                Y = sector_by_units.get(digits[-1])
        if X is None:
            X = cell[-2]
        if Y is None:
            Y = cell[-1]
        if X not in band_map or X == "_NOT_USED_":
            skipped.append((cell, f"unknown X={X}"))
            continue
        if Y not in sector_map:
            skipped.append((cell, f"unknown Y={Y}"))
            continue
        key = (site_new, ne_id, Y)
        if key not in seen_sector:
            seen_sector.add(key)
            sector_order.append(key)
        # First non-blank SiteName_Old wins; a sector's band-D rows often have
        # a blank _Old, which must not clobber the band-C row's real value (it
        # is the 3G-sheet lookup key for the CRb3/NSN_U2100 device).
        if site_old and key not in sector_site_old:
            sector_site_old[key] = site_old
        # SiteName_Old is a property of the whole physical site (identical for
        # every sector), but some CDDs leave it blank on every row of a given
        # sector. Keep a site-wide first-non-blank value so such a sector can
        # inherit it for the 3G/CRb3 lookup instead of falling to NOT_USED.
        site_key = (site_new, ne_id)
        if site_old and site_key not in site_old_by_site:
            site_old_by_site[site_key] = site_old
        ne_name = _get(r, "ne_name")
        if ne_name not in (None, "") and key not in sector_ne_name:
            sector_ne_name[key] = str(ne_name).strip()
        # Capture TxRxMode (New) per band so the L1800-band naming can prefer
        # the primary (L1800) cell's mode, falling back to any non-blank value.
        by_sector_txrx[key][X] = _get(r, "tx_rx_mode")
        try:
            e_tilt_val = float(e_tilt) if e_tilt not in (None, "-") else 0.0
        except (TypeError, ValueError):
            e_tilt_val = 0.0
        by_sector[key][X] = e_tilt_val

    rows = []
    # Index for the text (MML) feature, keyed by SiteName_New (== RET_input.txt
    # site line). Holds the {SiteName_New}_{Ne ID} DEVICENAME prefix and the
    # per-(sector_id, device position) RCU Tilt. Kept independent of the xlsx
    # RRU-Name prefix so the MML keeps using SiteName_New per CLAUDE.md.
    site_index = {}
    for key in sector_order:
        site_new, ne_id, Y = key
        sec = sector_map[Y]
        site_entry = site_index.setdefault(
            site_new, {"prefix": f"{site_new}_{ne_id}", "tilts": {}}
        )
        # Output 'Site Name(*)' = NEName_New from the CDD (falls back to
        # {site_new}_{ne_id} when that column is missing). The RRU/Device Name
        # prefix name is chosen by prefix_source (ne_name by default), so it is
        # the NEName_New (e.g. DTPLTN01N), not SiteName_New (e.g. DTPLTN01).
        ne_name_val = sector_ne_name.get(key)
        site_name = ne_name_val or f"{site_new}_{ne_id}"
        sources = {"ne_name": ne_name_val, "site_new": site_new}
        prefix_name = sources.get(prefix_source) or sources.get(prefix_fallback) or site_new
        sector_num = sec["sector_id"][1:]  # "S1" -> "1"
        site_old = sector_site_old.get(key) or site_old_by_site.get((site_new, ne_id))
        present = by_sector[key]  # bands found in CDD for this sector -> e_tilt

        # Resolve the L1800-band naming case (RET-Tool-logic.xlsx Cases 1-6)
        # once per sector: which token the Lb1 (_1) and CLb2 (_2) devices get,
        # from band presence + the sector TxRxMode (first match wins). CRb3
        # (band E) and Rb4 (NOT_USED) are unaffected by this.
        l1800_case = None
        if l1800_naming:
            txrx_map = by_sector_txrx.get(key, {})
            txrx = txrx_map.get(l1800_naming["primary_band"])
            if txrx in (None, ""):
                txrx = next((v for v in txrx_map.values() if v not in (None, "")), None)
            if txrx in (None, ""):
                txrx = l1800_naming.get("default_tx_rx_mode")
            has_l2100 = l1800_naming["l2100_band"] in present
            has_f2 = l1800_naming["f2_band"] in present
            for case in l1800_naming["cases"]:
                cond = case.get("when", {})
                if "l2100" in cond and cond["l2100"] != has_l2100:
                    continue
                if "f2" in cond and cond["f2"] != has_f2:
                    continue
                if "tx_rx_mode" in cond and cond["tx_rx_mode"] != txrx:
                    continue
                l1800_case = case
                break

        for pos, X in enumerate(DEVICE_ORDER):
            band = band_map[X]
            slot = band["slot_suffix"]
            band_token = band["band_token"]
            if X == "_NOT_USED_":
                rcu_tilt = 0
            elif X == tg_band:
                # Device No 3 (CRb3): presence + tilt come from the 3G sheet.
                if (site_old, sector_num) in three_g_tilts:
                    tg_tilt = three_g_tilts[(site_old, sector_num)]
                    rcu_tilt = round((tg_tilt or 0.0) * 10)
                else:
                    # No 3G tilt for this sector -> U2100_NOT_USED, whether the
                    # site is in 3G but lacks this sector (Case 8) or entirely
                    # absent from 3G. Tokens are configurable in mapping.json.
                    band_token = tg_partial_absent if site_old in three_g_sites else tg_absent
                    rcu_tilt = 0
            elif l1800_case and X == l1800_naming["device1_band"]:
                band_token = l1800_case["device1"]
                rcu_tilt = _scaled_tilt(
                    present, l1800_naming["tilt_band"], l1800_naming.get("tilt_fallback")
                )
            elif l1800_case and X == l1800_naming["device2_band"]:
                band_token = l1800_case["device2"]
                if "device2_tilt" in l1800_case:
                    rcu_tilt = l1800_case["device2_tilt"]
                elif "device2_tilt_band" in l1800_case:
                    rcu_tilt = _scaled_tilt(
                        present, l1800_case["device2_tilt_band"], l1800_naming["tilt_band"]
                    )
                else:
                    rcu_tilt = _scaled_tilt(
                        present, l1800_naming["tilt_band"], l1800_naming.get("tilt_fallback")
                    )
            else:
                rcu_tilt = round(by_sector[key].get(X, 0.0) * 10)
            rru_name = f"{prefix_name}_{ne_id}_{band_token}_{sec['sector_id']}{slot}"
            site_entry["tilts"][(sec["sector_id"], pos)] = rcu_tilt
            rows.append([
                site_name,
                rru_name,
                rru_cn,
                sec["rru_srn"],
                rru_sn,
                band["color"],
                rcu_tilt,
                rru_name,   # column H = column B (Device Name = RRU Name)
            ])

    return rows, skipped, len(sector_order), site_index


def _find_template_header_row(ws, headers, max_scan=20):
    """Find the column-header row in the template by matching known header tokens.

    Templates may have preamble rows above the header (e.g. an orange
    'Declaration' note), so the header is not always row 1. Returns the 1-based
    row index whose cells best match ``headers``; falls back to row 1.
    """
    targets = [_norm_header(h) for h in headers]
    max_col = ws.max_column or 1
    best_row, best_score = 1, 0
    for i in range(1, min(ws.max_row or 1, max_scan) + 1):
        cells = [_norm_header(ws.cell(row=i, column=c).value) for c in range(1, max_col + 1)]
        score = sum(1 for t in targets if t and any(t in cell for cell in cells))
        if score > best_score:
            best_score, best_row = score, i
    return best_row if best_score >= 2 else 1


def write_output(template_path, target_sheet, rows, output_path):
    """Write rows into a copy of the template, preserving its header rows and
    the styling of its first data row.

    The header row is detected (it may sit below preamble rows such as a
    Declaration note), so everything up to and including the header is kept and
    data is written starting on the next row.
    """
    out_wb = load_workbook(template_path)
    out_ws = out_wb[target_sheet]

    header_row = _find_template_header_row(out_ws, HEADERS)
    data_start = header_row + 1

    # Sample styling from the template's first data row (below the header);
    # if the template has no data rows, fall back to the header row.
    style_row = data_start if (out_ws.max_row or 0) >= data_start else header_row
    template_row_styles = []
    for col in range(1, out_ws.max_column + 1):
        cell = out_ws.cell(row=style_row, column=col)
        template_row_styles.append({
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "number_format": cell.number_format,
        })

    # Clear existing data rows only (keep preamble + header).
    if (out_ws.max_row or 0) >= data_start:
        out_ws.delete_rows(data_start, out_ws.max_row - data_start + 1)

    for i, values in enumerate(rows):
        out_row = data_start + i
        for col_idx, v in enumerate(values, start=1):
            c = out_ws.cell(row=out_row, column=col_idx, value=v)
            if col_idx - 1 < len(template_row_styles):
                s = template_row_styles[col_idx - 1]
                c.font = copy(s["font"])
                c.fill = copy(s["fill"])
                c.border = copy(s["border"])
                c.alignment = copy(s["alignment"])
                c.number_format = s["number_format"]

    out_wb.save(output_path)


# --------------------------------------------------------------------------
# RET_template.txt -> RET_output.txt (MML script) conversion.
#
# Rewrites three things in an "ADD RET / MOD RETTILT" MML template using the
# RET_input.txt serials and the CDD-derived rows (build_rows):
#   * DEVICENAME : replace the site-prefix tokens with {SiteName_New}_{Ne ID}.
#   * SERIALNO   : take the input serial matched by (CTRLSRN, RIGHT(serial,3)).
#   * TILT       : RCU Tilt of the same-DEVICENO ADD RET device, matched
#                  positionally (CTRLSRN -> sector, device order within sector).
# All rules/field names live in mapping.json -> text_config.
# --------------------------------------------------------------------------

# MML field regexes (DEVICENO/CTRLSRN allow an optional leading space, e.g.
# "DEVICENO= 0"; CTRLSRN is matched whole so it never collides with CTRLSN).
_RE_DEVICENO = re.compile(r"DEVICENO=\s*(\d+)")
_RE_DEVICENAME = re.compile(r'DEVICENAME="([^"]*)"')
_RE_CTRLSRN = re.compile(r"CTRLSRN=\s*(\d+)")
_RE_SERIALNO = re.compile(r'SERIALNO="([^"]*)"')
_RE_TILT = re.compile(r"TILT=\s*(-?\d+)")


def parse_ret_input(input_path, mapping):
    """Parse RET_input.txt -> (site, serials).

    Line 1 (first non-blank) is the site name. Each remaining line is split on
    whitespace; ``serials`` maps (CTRLSRN, RIGHT(serial, suffix_len)) -> full
    serial, the key used to rewrite each template SERIALNO.
    """
    with open(input_path, encoding="utf-8") as f:
        text = f.read()
    return parse_ret_input_text(text, mapping, source=input_path)


def parse_ret_input_text(text, mapping, source="pasted input"):
    """Same as parse_ret_input but on an in-memory string (e.g. pasted text)."""
    cfg = mapping.get("text_config", {}).get("input", {})
    srn_col = cfg.get("ctrlsrn_column", 1)
    ser_col = cfg.get("serial_column", 7)
    suf_len = cfg.get("serial_suffix_len", 3)

    lines = text.splitlines()

    site = None
    serials = {}
    for ln in lines:
        if not ln.strip():
            continue
        if site is None:
            site = ln.strip()
            continue
        parts = ln.split()
        if len(parts) <= max(srn_col, ser_col):
            continue
        srn = parts[srn_col].strip()
        serial = parts[ser_col].strip()
        serials[(srn, serial[-suf_len:])] = serial
    if site is None:
        raise ValueError("RET input is empty: %s" % source)
    return site, serials


def _site_tilt_index(site_index, site):
    """Return (new_prefix, tilt_by_sector_pos) for ``site`` from build_rows.

    ``site`` is the RET_input.txt site line, matched against SiteName_New (the
    key of ``site_index``) regardless of the xlsx RRU-Name prefix. ``new_prefix``
    is the {SiteName_New}_{Ne ID} DEVICENAME prefix; ``tilt_by_sector_pos`` maps
    (sector_id, position) -> RCU Tilt, position being the device's 0-based order
    within its sector (the template's per-sector device order).
    """
    entry = site_index.get(site)
    if entry is None:
        available = ", ".join(sorted(site_index))
        raise ValueError(
            "Site %r (from RET input) was not found in the CDD output.\n"
            "Sites available in the CDD: %s" % (site, available)
        )
    return entry["prefix"], entry["tilts"]


def build_text_output(template_path, input_path, cdd_path, sheet, mapping,
                      input_text=None):
    """Produce the rewritten MML text. Returns (text, warnings).

    Uses build_rows(cdd_path, sheet) for Ne ID and RCU Tilt, and parse_ret_input
    for the replacement serials. If ``input_text`` is given (e.g. pasted into the
    GUI), it is parsed instead of reading ``input_path``.
    """
    if input_text is not None and input_text.strip():
        site, serials = parse_ret_input_text(input_text, mapping)
    else:
        site, serials = parse_ret_input(input_path, mapping)
    _rows, _, _, site_index = build_rows(cdd_path, sheet, mapping)
    new_prefix, tilt_by_sector_pos = _site_tilt_index(site_index, site)

    srn_to_sector = {
        str(v["rru_srn"]): v["sector_id"] for v in mapping["sector_map"].values()
    }

    tcfg = mapping.get("text_config", {}).get("template", {})
    prefix_tokens = tcfg.get("prefix_token_count", 2)
    add_prefix = tcfg.get("add_line_prefix", "ADD RET")
    tilt_prefix = tcfg.get("tilt_line_prefix", "MOD RETTILT")
    suf_len = mapping.get("text_config", {}).get("input", {}).get("serial_suffix_len", 3)

    with open(template_path, encoding="utf-8") as f:
        tmpl_lines = f.readlines()

    warnings = []

    # Pass 1: DEVICENO -> RCU Tilt, using each ADD RET line's CTRLSRN (sector)
    # and its order within that sector (positional match).
    deviceno_tilt = {}
    add_pos = defaultdict(int)
    for ln in tmpl_lines:
        if not ln.lstrip().startswith(add_prefix):
            continue
        m_no, m_srn = _RE_DEVICENO.search(ln), _RE_CTRLSRN.search(ln)
        if not (m_no and m_srn):
            continue
        srn = m_srn.group(1)
        pos = add_pos[srn]
        add_pos[srn] += 1
        sector_id = srn_to_sector.get(srn)
        tilt = tilt_by_sector_pos.get((sector_id, pos))
        if tilt is None:
            warnings.append(
                "No RCU Tilt for DEVICENO=%s (CTRLSRN=%s, sector=%s, pos=%d)"
                % (m_no.group(1), srn, sector_id, pos)
            )
        deviceno_tilt[m_no.group(1)] = tilt

    # Pass 2: rewrite lines in place, preserving everything else verbatim.
    out_lines = []
    for ln in tmpl_lines:
        stripped = ln.lstrip()
        if stripped.startswith(add_prefix):
            def _sub_devicename(m):
                tokens = m.group(1).split("_")
                suffix = "_".join(tokens[prefix_tokens:])
                return 'DEVICENAME="%s"' % (new_prefix + "_" + suffix)

            ln = _RE_DEVICENAME.sub(_sub_devicename, ln)

            m_srn = _RE_CTRLSRN.search(ln)
            if m_srn:
                srn = m_srn.group(1)

                def _sub_serial(m):
                    suffix = m.group(1)[-suf_len:]
                    new = serials.get((srn, suffix))
                    if new is None:
                        warnings.append(
                            "No input serial for CTRLSRN=%s suffix=%s" % (srn, suffix)
                        )
                        return m.group(0)
                    return 'SERIALNO="%s"' % new

                ln = _RE_SERIALNO.sub(_sub_serial, ln)
            out_lines.append(ln)
        elif stripped.startswith(tilt_prefix):
            m_no = _RE_DEVICENO.search(ln)
            if m_no:
                tilt = deviceno_tilt.get(m_no.group(1))
                if tilt is not None:
                    ln = _RE_TILT.sub("TILT=%d" % int(round(float(tilt))), ln)
                else:
                    warnings.append("No tilt for MOD RETTILT DEVICENO=%s" % m_no.group(1))
            out_lines.append(ln)
        else:
            out_lines.append(ln)

    return "".join(out_lines), warnings


def write_text_output(template_path, input_path, cdd_path, sheet, mapping, output_path,
                      input_text=None):
    """build_text_output + write to ``output_path``. Returns warnings."""
    text, warnings = build_text_output(
        template_path, input_path, cdd_path, sheet, mapping, input_text=input_text
    )
    with open(output_path, "w", encoding="utf-8") as f:
        f.write(text)
    return warnings
