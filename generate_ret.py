"""Generate RETConfigWDTInternal_new.xlsx from CDD.xlsx using mapping.json."""
import json
from collections import defaultdict
from copy import copy

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

CDD_PATH = "CDD.xlsx"
TEMPLATE_PATH = "RETConfigWDTInternal.xlsx"
MAPPING_PATH = "mapping.json"
OUTPUT_PATH = "RETConfigWDTInternal_new.xlsx"


def main():
    with open(MAPPING_PATH, encoding="utf-8") as f:
        m = json.load(f)

    band_map = m["band_map"]
    sector_map = m["sector_map"]

    # 4 devices per sector: order matches positions 1..4 in the template.
    device_order = ["C", "D", "E", "_NOT_USED_"]

    three_g = m.get("three_g", {})
    tg_band = three_g.get("governs_band")            # e.g. "E"
    tg_absent = three_g.get("absent_band_token", "NOT_USED")

    # Read CDD: collect (site_new, ne_id, Y) -> {X -> e_tilt}
    cdd_wb = load_workbook(CDD_PATH, data_only=True)
    cdd_ws = cdd_wb[m["source"]["sheet"]]

    rows = list(cdd_ws.iter_rows(min_row=2, values_only=True))

    # (site_old, sector number) -> New E-Tilt from the 3G Installation Design sheet.
    three_g_tilts = {}
    if three_g and three_g.get("sheet") in cdd_wb.sheetnames:
        tg_ws = cdd_wb[three_g["sheet"]]
        tg_col = ord(three_g.get("logical_sector_column", "B").upper()) - ord("A")
        tg_tilt_col = ord(three_g.get("etilt_column", "F").upper()) - ord("A")
        tg_sep = three_g.get("separator", "-")
        for r in tg_ws.iter_rows(min_row=2, values_only=True):
            if len(r) <= tg_col or r[tg_col] is None:
                continue
            so, _, num = str(r[tg_col]).partition(tg_sep)
            if not num:
                continue
            tg_key = (so.strip(), num.strip())
            et = r[tg_tilt_col] if len(r) > tg_tilt_col else None
            try:
                et_val = float(et) if et not in (None, "-", "") else None
            except (TypeError, ValueError):
                et_val = None
            if tg_key not in three_g_tilts or (three_g_tilts[tg_key] is None and et_val is not None):
                three_g_tilts[tg_key] = et_val

    # by_sector[(site_new, ne_id, Y)] = {X: e_tilt}
    by_sector = defaultdict(dict)
    sector_site_old = {}  # (site_new, ne_id, Y) -> site_old
    # Preserve insertion order of (site, sector) to keep output stable.
    sector_order = []
    seen_sector = set()
    skipped = []

    for r in rows:
        if not r or r[6] is None:
            continue
        site_old, site_new, main_site, ne_name, enb, ne_id, cell, e_tilt, m_old, m_new = r[:10]
        cell = str(cell)
        if len(cell) < 2:
            skipped.append((cell, "too short"))
            continue
        X = cell[-2]
        Y = cell[-1]
        if X not in band_map or X == "_NOT_USED_":
            skipped.append((cell, f"unknown X={X}"))
            continue
        if Y not in sector_map:
            skipped.append((cell, f"unknown Y={Y}"))
            continue
        key = (site_new, ne_id, Y)
        if key not in seen_sector:
            seen_sector.add(key)
            sector_order.append(key)
        sector_site_old[key] = site_old
        try:
            e_tilt_val = float(e_tilt) if e_tilt is not None and e_tilt != "-" else 0.0
        except (TypeError, ValueError):
            e_tilt_val = 0.0
        by_sector[key][X] = e_tilt_val

    # Open template, copy header + first data row styling, then rewrite data rows.
    out_wb = load_workbook(TEMPLATE_PATH)
    out_ws = out_wb[m["target"]["sheet"]]

    # Capture style from row 2 (first data row) to replicate.
    template_row_styles = []
    for col in range(1, out_ws.max_column + 1):
        cell = out_ws.cell(row=2, column=col)
        template_row_styles.append({
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "number_format": cell.number_format,
        })

    # Clear existing data rows (keep header at row 1).
    if out_ws.max_row >= 2:
        out_ws.delete_rows(2, out_ws.max_row - 1)

    out_row = 2
    for key in sector_order:
        site_new, ne_id, Y = key
        sec = sector_map[Y]
        site_name = f"{site_new}_{ne_id}"
        sector_num = sec["sector_id"][1:]  # "S1" -> "1"
        site_old = sector_site_old.get(key)
        for X in device_order:
            band = band_map[X]
            band_token = band["band_token"]
            slot = band["slot_suffix"]  # "_1" or "_2"
            color = band["color"]
            if X == "_NOT_USED_":
                rcu_tilt = 0
            elif X == tg_band:
                # Device No 3: presence + tilt come from the 3G sheet.
                if (site_old, sector_num) in three_g_tilts:
                    tg_tilt = three_g_tilts[(site_old, sector_num)]
                    rcu_tilt = round((tg_tilt or 0.0) * 10)
                else:
                    band_token = tg_absent
                    rcu_tilt = 0
            else:
                e_tilt_val = by_sector[key].get(X, 0.0)
                rcu_tilt = round(e_tilt_val * 10)
            rru_name = f"{site_new}_{ne_id}_{band_token}_{sec['sector_id']}{slot}"
            values = [
                site_name,
                rru_name,
                m["constants"]["rru_cn"],
                sec["rru_srn"],
                m["constants"]["rru_sn"],
                color,
                rcu_tilt,
                rru_name,
            ]
            for col_idx, v in enumerate(values, start=1):
                c = out_ws.cell(row=out_row, column=col_idx, value=v)
                s = template_row_styles[col_idx - 1]
                c.font = copy(s["font"])
                c.fill = copy(s["fill"])
                c.border = copy(s["border"])
                c.alignment = copy(s["alignment"])
                c.number_format = s["number_format"]
            out_row += 1

    out_wb.save(OUTPUT_PATH)
    total_rows = out_row - 2
    print(f"Wrote {OUTPUT_PATH}: {total_rows} data rows across {len(sector_order)} sectors.")
    if skipped:
        print(f"Skipped {len(skipped)} CDD rows:")
        for s in skipped:
            print(" -", s)


if __name__ == "__main__":
    main()
